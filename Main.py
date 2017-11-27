import numpy as np
import cv2
import math
import time
from skimage import feature
from sklearn.neighbors import KNeighborsClassifier
from openpyxl import load_workbook

from tkinter.filedialog import askopenfilename

def rgb_para_cinza(img):
	'''
	Converte a imagem de RGB para escala de cinza. Eu uso no lugar da funcao do opencv porque ela da pesos diferentes às cores.
	'''
	b,g,r = cv2.split(img)
	return (b // 3 + g // 3 + r // 3)

def diferenca(imgA, imgB):
	'''
	Checa a diferenca entre duas imagens
	'''
	diff = np.zeros((rows, columns))
	for r in range(rows):
		for c in range(columns):
			if imgA[r, c, 0] != imgB[r, c, 0] or imgA[r, c, 1] != imgB[r, c, 1] or imgA[r, c, 2] != imgB[r, c, 2]:
				diff[r, c] = 255
	return diff

def gerar_blocos(img):
	'''
	Cria os blocos para calcular o histograma de cada um. Os blocos sao representados apenas pela linha e coluna do pixel mais acima e a esquerda do bloco.
	'''
	r = 0
	while r < rows - b + 1:
		c = 0
		while c < columns - b + 1:
			yield (r, c)
			c += espaco
		r += espaco

def gerar_histograma(lbps, lbpMax):
	'''
	Gera um histograma de um conjunto de LBPs
	'''
	histograma = np.zeros(lbpMax + 1, dtype=int)
	for lbp in lbps:
		histograma[int(lbp)] += 1
	return histograma

def num_blocos_linha():
	return (columns - b) // espaco + 1

def num_blocos_coluna():
	return (rows - b) // espaco + 1

def num_blocos():
	return num_blocos_linha() * num_blocos_coluna() 

def gerar_histograma_blocos(img, lbps, lbpMax):
	'''
	Gera histogramas de todos os blocos nos quais as imagens foram divididas
	'''
	saida = np.zeros((num_blocos(), lbpMax + 1))
	i = 0
	for (r, c) in gerar_blocos(img):
		lista = np.zeros(b*b, dtype=int)
		j = 0
		for x in range(r, r+b):
			for y in range(c, c+b):
				lista[j] = lbps[x, y]
				j += 1
		saida[i] = gerar_histograma(lista, lbpMax)
		i += 1
	return saida

def knn(vetores, pixelsBloco):
	'''
	Obtem o vizinho mais proximo (que nao seja ele mesmo) de cada histograma
	'''
	saida = np.zeros(len(vetores), dtype=int)

	neigh = KNeighborsClassifier(n_neighbors=1)
	indices = np.array([x for x in range(len(vetores))])
	neigh.fit(vetores, indices)
	for i in range(len(vetores)):
		vetor = np.copy(vetores[i])
		vetores[i] = np.array([1000 for _ in vetores[i]])
		saida[i] = neigh.predict([vetor])
		vetores[i] = vetor
	
	return saida

def identificar2(vA, vB, vC):
	'''
	Identifica os blocos copiados, diferenciando entre original e copia, embora nao possa dizer qual e qual
	'''
	saidaA = np.zeros(len(vA), dtype=bool)
	saidaB = np.zeros(len(vA), dtype=bool)
	saidaC = np.zeros(len(vA), dtype=bool)
	saidaD = np.zeros(len(vA), dtype=bool)
	
	for i in range(len(vA)):
		if vA[i] == vB[i] or vB[i] == vC[i] or vC[i] == vA[i]:
			if not saidaB[i]:
				saidaA[i] = True
				if vA[i] == vB[i]:
					saidaB[vA[i]] = True
				else:
					saidaB[vC[i]] = True
		if vA[i] == vB[i] and vA[i] == vC[i]:
			if not saidaD[i]:
				saidaC[i] = True
				saidaD[vA[i]] = True
	return (saidaA, saidaB, saidaC, saidaD)

def indice_bloco(r, c):
	'''
	Obtem o indice de um bloco no array de blocos
	'''
	return r * num_blocos_linha() // espaco + c // espaco
	
def editar(img, blocosA, blocosB):
	'''
	Deixa um conjunto de blocos na imagem azul e outro verde
	'''
	saida = np.copy(img)
	for (r, c) in gerar_blocos(img):
		if blocosA[indice_bloco(r, c)]:
			saida[r:r+b, c:c+b] = [255, 0, 0]
		if blocosB[indice_bloco(r, c)]:
			saida[r:r+b, c:c+b] = [0, 255, 0]
	return saida

def editar_blocos(img, blocosA, blocosB):
	'''
	Gera uma figura com os blocos atribuidos a cada grupo 
	'''
	i = 0
	saida = np.zeros((num_blocos_coluna(), num_blocos_linha(), 3))
	for (r, c) in gerar_blocos(img):
		if blocosA[i]:
			saida[r // espaco, c // espaco] = [255, 0, 0]
		if blocosB[i]:
			saida[r // espaco, c // espaco] = [0, 255, 0]
		i += 1
	return saida

def checar_vizinhos(blocos, row, col, largura):
	'''
	Checa se os vizinhos de um pixel sao do grupo dele
	'''
	for r in range(row // espaco - largura + 1, row // espaco + largura):
		for c in range(col // espaco - largura + 1, col // espaco + largura):
			if r < 0 or r >= num_blocos_coluna() or c < 0 and c >= num_blocos_coluna():
				return False
			if not blocos[indice_bloco(r, c)]:
				return False
	return True

def editar_filtrado(img, blocosA, blocosB):
	'''
	Gera uma imagem filtrada, com apenas os blocos cujos vizinhos sao do mesmo grupo azul ou verde. Tambem da um veredito se a imagem tem copias ou nao.
	'''
	saida = np.copy(img)
	temAzul = False
	temVerde = False
	for (r, c) in gerar_blocos(img):
		if blocosA[indice_bloco(r, c)] and checar_vizinhos(blocosA, r, c, 3):
			saida[r:r+b, c:c+b] = [255, 0, 0]
			temAzul = True
		if blocosB[indice_bloco(r, c)] and checar_vizinhos(blocosB, r, c, 3):
			saida[r:r+b, c:c+b] = [0, 255, 0]
			temVerde = True
	return (saida, temAzul and temVerde)

def precisao_e_recall(teste):
	truePosA = 0
	falsePosA = 0
	falseNegA = 0
	truePosB = 0
	falsePosB = 0
	falseNegB = 0
	for r in range(rows):
		for c in range(columns):
			if teste[r, c, 0] == 255 and teste[r, c, 1] == 0 and teste[r, c, 2] == 0:
				if orig[r, c] == 255:
					truePosA += 1
					falseNegB += 1
				else:
					falsePosA += 1
			elif teste[r, c, 0] == 0 and teste[r, c, 1] == 255 and teste[r, c, 2] == 0:
				if orig[r, c] == 255:
					falseNegA += 1
					truePosB += 1
				else:
					falsePosB += 1
			elif orig[r, c] == 255:
				falseNegA += 1
				falseNegB += 1
	saidaA = (truePosA, falsePosA, falseNegA)
	saidaB = (truePosB, falsePosB, falseNegB)
	if truePosA > truePosB:
		return saidaA
	elif truePosA < truePosB:
		return saidaB
	elif falsePosA < falsePosB:
		return saidaA
	else:
		return saidaB

nomeOrig = askopenfilename(filetypes=[("all files","*"),("Bitmap Files","*.bmp; *.dib"), ("JPEG", "*.jpg; *.jpe; *.jpeg; *.jfif"), ("PNG", "*.png"), ("TIFF", "*.tiff; *.tif")])
nomeMod = askopenfilename(filetypes=[("all files","*"),("Bitmap Files","*.bmp; *.dib"), ("JPEG", "*.jpg; *.jpe; *.jpeg; *.jfif"), ("PNG", "*.png"), ("TIFF", "*.tiff; *.tif")])
nomeOrig = nomeOrig[nomeOrig.find("/Imagens/") + 1:]
nomeMod = nomeMod[nomeMod.find("/Imagens/") + 1:]

tempo = time.time()
orig = cv2.imread(nomeOrig)
img = cv2.imread(nomeMod)
print(nomeOrig, nomeMod)
cinza = rgb_para_cinza(img)
b = 10
espaco = 1
(rows, columns) = cinza.shape
print("Experimento com b = " + str(b) + ", espaco = " + str(espaco))

#Calculando a diferenca entre o original e a copia
tempoEtapa = time.time()
orig = diferenca(orig, img)
print("Tempo de calculo da diferenca real: " + str(time.time() - tempoEtapa))

#Calculo de LBPs
tempoEtapa = time.time()
lbpA = feature.local_binary_pattern(cinza, 8, 1, method="nri_uniform")
lbpB = feature.local_binary_pattern(cinza, 12, 2, method="nri_uniform")
lbpC = feature.local_binary_pattern(cinza, 16, 2, method="uniform")
print("Tempo de calculo do LBP: " + str(time.time() - tempoEtapa))

#Geracao de histogramas
tempoEtapa = time.time()
histogramaA = gerar_histograma_blocos(cinza, lbpA, 8 * (8-1) + 2)
histogramaB = gerar_histograma_blocos(cinza, lbpB, 12 * (12-1) + 2)
histogramaC = gerar_histograma_blocos(cinza, lbpC, 16 + 1)
print("Tempo de geracao dos histogramas: " + str(time.time() - tempoEtapa))
print("Num. de histogramas: " + str(num_blocos()))

#Obtencao dos blocos mais parecidos
tempoEtapa = time.time()
proximosA = knn(histogramaA, b*b)
print("Tempo de calculo do KNN A: " + str(time.time() - tempoEtapa))
tempoKnnB = time.time()
proximosB = knn(histogramaB, b*b)
print("Tempo de calculo do KNN B: " + str(time.time() - tempoKnnB))
tempoKnnC = time.time()
proximosC = knn(histogramaC, b*b)
print("Tempo de calculo do KNN C: " + str(time.time() - tempoKnnC))
print("Tempo de calculo do KNN: " + str(time.time() - tempoEtapa))

#Obtencao das regioes original e copiada
tempoEtapa = time.time()
(idA2, idB2, idA3, idB3) = identificar2(proximosA, proximosB, proximosC)
print("Tempo de identificacao de copias: " + str(time.time() - tempoEtapa))

#Geracao de imagens
tempoEtapa = time.time()
img2 = editar(np.zeros(img.shape), idA2, idB2)
img3 = editar(np.zeros(img.shape), idA3, idB3)
img2B = editar_blocos(np.zeros(img.shape), idA2, idB2)
img3B = editar_blocos(np.zeros(img.shape), idA3, idB3)
(img2C, copia2) = editar_filtrado(np.zeros(img.shape), idA2, idB2)
(img3C, copia3) = editar_filtrado(np.zeros(img.shape), idA3, idB3)
print("Tempo de edicao: " + str(time.time() - tempoEtapa))

#Calculo do desempenho
tempoEtapa = time.time()
print("COM 2 COINCIDENCIAS:")
(truePos, falsePos2, falseNeg) = precisao_e_recall(img2C)
print("Verdadeiros positivos: " + str(truePos))
print("Falsos positivos: " + str(falsePos2))
print("Falsos negativos: " + str(falseNeg))
if (truePos + falseNeg):
	recall2 = truePos / (truePos + falseNeg)
	print("Recall = " + str(recall2))
precisao2 = 0
if (truePos + falsePos2):
	precisao2 = truePos / (truePos + falsePos2)
	print("Precisao = " + str(precisao2))

print("COM 3 COINCIDENCIAS:")
(truePos, falsePos3, falseNeg) = precisao_e_recall(img3C)
print("Verdadeiros positivos: " + str(truePos))
print("Falsos positivos: " + str(falsePos3))
print("Falsos negativos: " + str(falseNeg))
if (truePos + falseNeg):
	recall3 = truePos / (truePos + falseNeg)
	print("Recall = " + str(recall3))
precisao3 = 0
if (truePos + falsePos3):
	precisao3 = truePos / (truePos + falsePos3)
	print("Precisao = " + str(precisao3))
print("Tempo para calcular desempenho: " + str(time.time() - tempoEtapa))
	
print("Tempo de Execucao: " + str(time.time() - tempo))

print("Veredito de 2 coincidencias: ", end="")
if (copia2):
	print("Tem copia")
else:
	print("Nao tem copia")
print("Veredito de 3 coincidencias: ", end="")
if (copia3):
	print("Tem copia")
else:
	print("Nao tem copia")

#Exibicao das imagens
cv2.imshow("Copia e Original - min. 2", img2)
cv2.imshow("Copia e Original - min. 3", img3)
cv2.imshow("Classificacao dos blocos - min. 2", img2B)
cv2.imshow("Classificacao dos blocos - min. 3", img3B)
cv2.imshow("Filtragem de ruido - min. 2", img2C)
cv2.imshow("Filtragem de ruido - min. 3", img3C)
cv2.imshow("Diferenca real", orig)
cv2.waitKey(0)
cv2.destroyAllWindows()

#Gravacao dos resultados
xlsx = load_workbook("Resultados.xlsx")
if nomeOrig != nomeMod:
	numTeste = int(nomeMod.replace("Imagens/Modificado ", "")[:-4])
	cv2.imwrite(nomeMod.replace("Modificado", "Diferenca2Coincidencias"), img2)
	cv2.imwrite(nomeMod.replace("Modificado", "Diferenca3Coincidencias"), img3)
	cv2.imwrite(nomeMod.replace("Modificado", "Blocos2Coincidencias"), img2B)
	cv2.imwrite(nomeMod.replace("Modificado", "Blocos3Coincidencias"), img3B)
	cv2.imwrite(nomeMod.replace("Modificado", "Filtrado2Coincidencias"), img2C)
	cv2.imwrite(nomeMod.replace("Modificado", "Filtrado3Coincidencias"), img3C)
	cv2.imwrite(nomeMod.replace("Modificado", "Diferenca Real"), orig)

	planilha = xlsx["Modificados"]
	planilha.cell(row = numTeste + 2, column = 1, value = numTeste)
	planilha.cell(row = numTeste + 2, column = 2, value = recall2)
	planilha.cell(row = numTeste + 2, column = 3, value = precisao2)
	planilha.cell(row = numTeste + 2, column = 4, value = copia2)
	planilha.cell(row = numTeste + 2, column = 5, value = recall3)
	planilha.cell(row = numTeste + 2, column = 6, value = precisao3)
	planilha.cell(row = numTeste + 2, column = 7, value = copia3)
else:
	numTeste = int(nomeMod.replace("Imagens/Original ", "")[:-4])
	cv2.imwrite(nomeOrig.replace("Original", "Controle2Coincidencias"), img2B)
	cv2.imwrite(nomeOrig.replace("Original", "Controle3Coincidencias"), img3B)
	cv2.imwrite(nomeOrig.replace("Original", "ControleFiltrado2Coincidencias"), img2C)
	cv2.imwrite(nomeOrig.replace("Original", "ControleFiltrado3Coincidencias"), img3C)
	
	planilha = xlsx["Originais"]
	planilha.cell(row = numTeste + 2, column = 1, value = numTeste)
	planilha.cell(row = numTeste + 2, column = 2, value = falsePos2)
	planilha.cell(row = numTeste + 2, column = 3, value = copia2)
	planilha.cell(row = numTeste + 2, column = 4, value = falsePos3)
	planilha.cell(row = numTeste + 2, column = 5, value = copia3)
xlsx.save("Resultados.xlsx")